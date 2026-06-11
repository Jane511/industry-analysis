"""Canonical ABS/manual public-data loaders and parsers."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.ptrs_reconstruction import PTRS_MODEL_NOTE
from src.utils import normalise_text

def load_rba_cash_rate(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, skiprows=[0,2,3,4,5,6,7,8,9,10], encoding='utf-8-sig')
    df = df.rename(columns={'Title':'date'})
    df['date'] = pd.to_datetime(df['date'], format='%d-%b-%Y', errors='coerce')
    df['Cash Rate Target'] = pd.to_numeric(df['Cash Rate Target'], errors='coerce')
    return df[['date','Cash Rate Target']].dropna()

def parse_abs_timeseries_xlsx(path: Path, measure_name: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name='Data1', header=None)
    dates = pd.to_datetime(df.iloc[9:,0], format='mixed', errors='coerce')
    out = []
    for j in range(1, df.shape[1]):
        header = df.iloc[0, j]
        if pd.isna(header):
            continue
        industry = str(header).split(';')[2].strip() if ';' in str(header) else str(header)
        values = pd.to_numeric(df.iloc[9:,j], errors='coerce')
        tmp = pd.DataFrame({'date': dates, 'value': values}).dropna(subset=['date'])
        tmp['industry'] = industry
        tmp['measure'] = measure_name
        out.append(tmp)
    return pd.concat(out, ignore_index=True)

def parse_labour_force(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name='Data1', header=None)
    dates = pd.to_datetime(df.iloc[9:,0], format='mixed', errors='coerce')
    out = []
    for j in range(1, df.shape[1]):
        industry_raw = df.iloc[0, j]
        series_type = df.iloc[2, j]
        if pd.isna(industry_raw) or pd.isna(series_type):
            continue
        industry = str(industry_raw).split(';')[0].strip()
        values = pd.to_numeric(df.iloc[9:,j], errors='coerce')
        tmp = pd.DataFrame({'date': dates, 'value': values}).dropna(subset=['date'])
        tmp['industry'] = industry
        tmp['series_type'] = series_type
        out.append(tmp)
    return pd.concat(out, ignore_index=True)

def _read_abs_data1(path: Path) -> pd.DataFrame:
    """Return the canonical ABS time-series dataframe with header rows intact."""
    return pd.read_excel(path, sheet_name="Data1", header=None)


def _abs_dates(df: pd.DataFrame) -> pd.Series:
    return pd.to_datetime(df.iloc[9:, 0], format="mixed", errors="coerce")


def _abs_long_series(df: pd.DataFrame, header_part_index: int = 0) -> pd.DataFrame:
    """Melt an ABS Data1 sheet into a long ``date | series_label | value`` frame.

    Each ABS series header is a semicolon-delimited label such as
    "Index Numbers ; All groups CPI ; Australia ;". ``header_part_index``
    selects which segment becomes ``series_label``.
    """
    dates = _abs_dates(df)
    out = []
    for j in range(1, df.shape[1]):
        header = df.iloc[0, j]
        if pd.isna(header):
            continue
        parts = [p.strip() for p in str(header).split(";") if p.strip()]
        if not parts:
            continue
        label = parts[header_part_index] if header_part_index < len(parts) else parts[-1]
        values = pd.to_numeric(df.iloc[9:, j], errors="coerce")
        tmp = pd.DataFrame({"date": dates, "value": values}).dropna(subset=["date"])
        tmp["series_label"] = label
        out.append(tmp)
    if not out:
        return pd.DataFrame(columns=["date", "value", "series_label"])
    return pd.concat(out, ignore_index=True)


def _yoy_qoq(series: pd.Series) -> tuple[float, float]:
    """Compute YoY (4-period) and QoQ (1-period) % change for an index series."""
    series = series.dropna()
    if len(series) < 2:
        return float("nan"), float("nan")
    latest = float(series.iloc[-1])
    qoq = (latest / float(series.iloc[-2]) - 1) * 100 if series.iloc[-2] else float("nan")
    if len(series) >= 5:
        yoy = (latest / float(series.iloc[-5]) - 1) * 100 if series.iloc[-5] else float("nan")
    else:
        yoy = float("nan")
    return yoy, qoq


def _empty_cpi_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "as_of_date",
            "period_label",
            "all_groups_index",
            "all_groups_yoy_pct",
            "all_groups_qoq_pct",
            "housing_index",
            "housing_yoy_pct",
            "housing_qoq_pct",
            "transport_index",
            "transport_yoy_pct",
            "transport_qoq_pct",
            "food_index",
            "food_yoy_pct",
            "food_qoq_pct",
            "source_note",
        ]
    )


def parse_cpi(all_groups_path: Path, subgroups_path: Path) -> pd.DataFrame:
    """Parse ABS Cat. 6401.0 CPI quarterly data.

    Returns one row per quarter with all-groups + housing/transport/food
    subgroup index levels and their YoY/QoQ % change. Both files are required
    inputs; if either is absent an empty DataFrame with the canonical schema
    is returned (graceful-degradation contract).
    """
    if not all_groups_path.exists() or not subgroups_path.exists():
        return _empty_cpi_frame()

    all_groups = _abs_long_series(_read_abs_data1(all_groups_path))
    subgroups = _abs_long_series(_read_abs_data1(subgroups_path), header_part_index=1)

    if all_groups.empty:
        return _empty_cpi_frame()

    # All-groups national series — first non-null series is "All groups CPI".
    ag_series = (
        all_groups[all_groups["series_label"].str.contains("All groups", case=False, na=False)]
        .sort_values("date")
        .drop_duplicates(subset=["date"], keep="last")
    )
    if ag_series.empty:
        ag_series = all_groups.sort_values("date").groupby("date", as_index=False)["value"].mean()
        ag_series["series_label"] = "All groups CPI"

    rows = []
    quarters = sorted(ag_series["date"].unique())[-8:]
    label_to_field = {
        "Housing": "housing",
        "Transport": "transport",
        "Food and non-alcoholic beverages": "food",
    }

    for q in quarters:
        as_iso = pd.Timestamp(q).date().isoformat()
        period_label = pd.Timestamp(q).strftime("%YQ%q") if hasattr(pd.Timestamp(q), "quarter") else as_iso
        period_label = f"{pd.Timestamp(q).year}Q{pd.Timestamp(q).quarter}"

        ag_until = ag_series[ag_series["date"] <= q].sort_values("date")["value"]
        ag_yoy, ag_qoq = _yoy_qoq(ag_until)
        ag_latest = float(ag_until.iloc[-1]) if not ag_until.empty else float("nan")

        row = {
            "as_of_date": as_iso,
            "period_label": period_label,
            "all_groups_index": ag_latest,
            "all_groups_yoy_pct": round(ag_yoy, 2) if pd.notna(ag_yoy) else float("nan"),
            "all_groups_qoq_pct": round(ag_qoq, 2) if pd.notna(ag_qoq) else float("nan"),
        }
        for label, field in label_to_field.items():
            sg = subgroups[
                subgroups["series_label"].str.contains(label, case=False, na=False)
            ]
            sg_until = sg[sg["date"] <= q].sort_values("date")["value"]
            sg_yoy, sg_qoq = _yoy_qoq(sg_until)
            sg_latest = float(sg_until.iloc[-1]) if not sg_until.empty else float("nan")
            row[f"{field}_index"] = sg_latest
            row[f"{field}_yoy_pct"] = round(sg_yoy, 2) if pd.notna(sg_yoy) else float("nan")
            row[f"{field}_qoq_pct"] = round(sg_qoq, 2) if pd.notna(sg_qoq) else float("nan")
        row["source_note"] = (
            f"ABS Cat. 6401.0 CPI ({all_groups_path.name}, {subgroups_path.name})"
        )
        rows.append(row)
    return pd.DataFrame(rows)


def _empty_ppi_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "as_of_date",
            "anzsic_division_code",
            "industry",
            "ppi_index",
            "ppi_yoy_pct",
            "ppi_qoq_pct",
            "source_note",
        ]
    )


def parse_ppi(manufacturing_path: Path, construction_path: Path) -> pd.DataFrame:
    """Parse ABS Cat. 6427.0 Producer Price Index by industry.

    Maps the manufacturing table to ANZSIC division C and the construction
    table to division E. One row per (quarter, division). Returns the canonical
    empty frame if either file is missing.
    """
    if not manufacturing_path.exists() or not construction_path.exists():
        return _empty_ppi_frame()

    sources = [
        ("C", "Manufacturing", manufacturing_path),
        ("E", "Construction", construction_path),
    ]
    frames = []
    for code, name, path in sources:
        long_df = _abs_long_series(_read_abs_data1(path))
        if long_df.empty:
            continue
        # Aggregate across all sub-series to a single index level per quarter
        # (the headline industry index is the first/total series in the table).
        primary_label = long_df["series_label"].iloc[0]
        primary = long_df[long_df["series_label"] == primary_label].sort_values("date")
        for q in sorted(primary["date"].unique())[-8:]:
            until = primary[primary["date"] <= q]["value"]
            yoy, qoq = _yoy_qoq(until)
            latest = float(until.iloc[-1]) if not until.empty else float("nan")
            frames.append(
                {
                    "as_of_date": pd.Timestamp(q).date().isoformat(),
                    "anzsic_division_code": code,
                    "industry": name,
                    "ppi_index": latest,
                    "ppi_yoy_pct": round(yoy, 2) if pd.notna(yoy) else float("nan"),
                    "ppi_qoq_pct": round(qoq, 2) if pd.notna(qoq) else float("nan"),
                    "source_note": f"ABS Cat. 6427.0 PPI ({path.name})",
                }
            )
    if not frames:
        return _empty_ppi_frame()
    return pd.DataFrame(frames)


def _empty_dwelling_approvals_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "as_of_date",
            "dwelling_type",
            "approvals_count",
            "value_aud_thousands",
            "yoy_pct_change",
            "trend_3m_yoy_pct",
            "source_note",
        ]
    )


def parse_dwelling_approvals(path: Path) -> pd.DataFrame:
    """Parse ABS Cat. 8752.0 dwelling-unit residential building approvals.

    Returns one row per (latest-month, dwelling_type). Three dwelling_type
    values are emitted: ``houses``, ``units_apartments``, ``total``.
    Closes the residential-approvals gap that left the RES row of
    property_market_overlays as a placeholder.
    """
    if not path.exists():
        return _empty_dwelling_approvals_frame()

    long_df = _abs_long_series(_read_abs_data1(path))
    if long_df.empty:
        return _empty_dwelling_approvals_frame()

    # ABS publishes "Houses", "Units (other than houses)", and "Total" series
    # — case-insensitive match on the series label.
    label_to_type = {
        "houses": "houses",
        "other residential": "units_apartments",
        "units other than houses": "units_apartments",
        "total dwellings": "total",
        "total residential": "total",
    }

    rows: list[dict] = []
    latest_date = long_df["date"].max()
    cutoff_3m = latest_date - pd.DateOffset(months=3)

    for series_label, group in long_df.groupby("series_label"):
        key = series_label.lower()
        dwelling_type = next(
            (label_to_type[k] for k in label_to_type if k in key),
            None,
        )
        if dwelling_type is None:
            continue
        ordered = group.sort_values("date").dropna(subset=["value"])
        if ordered.empty:
            continue
        latest = ordered.iloc[-1]
        prev_year = ordered[ordered["date"] <= latest["date"] - pd.DateOffset(years=1)]
        prev = prev_year.iloc[-1] if not prev_year.empty else None
        recent_3m = ordered[ordered["date"] > cutoff_3m]
        prior_3m_window = ordered[
            (ordered["date"] <= cutoff_3m)
            & (ordered["date"] > cutoff_3m - pd.DateOffset(years=1))
        ]
        prior_3m_match = prior_3m_window.tail(len(recent_3m))

        yoy = (
            (float(latest["value"]) / float(prev["value"]) - 1) * 100
            if prev is not None and prev["value"]
            else float("nan")
        )
        trend_yoy = (
            (recent_3m["value"].mean() / prior_3m_match["value"].mean() - 1) * 100
            if not prior_3m_match.empty and prior_3m_match["value"].mean()
            else float("nan")
        )
        rows.append(
            {
                "as_of_date": pd.Timestamp(latest["date"]).date().isoformat(),
                "dwelling_type": dwelling_type,
                "approvals_count": float(latest["value"]),
                "value_aud_thousands": float("nan"),
                "yoy_pct_change": round(yoy, 2) if pd.notna(yoy) else float("nan"),
                "trend_3m_yoy_pct": round(trend_yoy, 2) if pd.notna(trend_yoy) else float("nan"),
                "source_note": f"ABS Cat. 8752.0 dwelling approvals ({path.name})",
            }
        )

    if not rows:
        return _empty_dwelling_approvals_frame()

    # Merge duplicate dwelling_types (e.g. multiple "total" series) by mean.
    out = pd.DataFrame(rows)
    out = (
        out.groupby(["as_of_date", "dwelling_type"], as_index=False)
        .agg(
            approvals_count=("approvals_count", "sum"),
            value_aud_thousands=("value_aud_thousands", "sum"),
            yoy_pct_change=("yoy_pct_change", "mean"),
            trend_3m_yoy_pct=("trend_3m_yoy_pct", "mean"),
            source_note=("source_note", "first"),
        )
    )
    return out


def parse_building_approvals(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name='Data1', header=None)
    dates = pd.to_datetime(df.iloc[9:,0], format='mixed', errors='coerce')
    out = []
    for j in range(1, df.shape[1]):
        header = df.iloc[0, j]
        if pd.isna(header):
            continue
        parts = [p.strip() for p in str(header).split(';') if p.strip()]
        if len(parts) < 3:
            continue
        values = pd.to_numeric(df.iloc[9:,j], errors='coerce')
        tmp = pd.DataFrame({'date': dates, 'value': values}).dropna(subset=['date'])
        tmp['sector_group'] = parts[1]
        tmp['building_type'] = parts[2]
        out.append(tmp)
    return pd.concat(out, ignore_index=True)

def parse_australian_industry_totals(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name='Table_1', header=None)
    df.columns = ['label','employment_000','wages_m','sales_m','total_income_m','total_expenses_m','op_profit_before_tax_m','ebitda_m','industry_value_added_m']
    records = []
    for i, row in df.iterrows():
        label = row['label']
        if isinstance(label, str) and label.lower().startswith('total '):
            for k in range(1,4):
                year_row = df.iloc[i+k]
                records.append({
                    'sector': label.replace('Total ','').strip(),
                    'year': str(year_row['label']),
                    'employment_000': year_row['employment_000'],
                    'wages_m': year_row['wages_m'],
                    'sales_m': year_row['sales_m'],
                    'op_profit_before_tax_m': year_row['op_profit_before_tax_m'],
                    'ebitda_m': year_row['ebitda_m'],
                    'industry_value_added_m': year_row['industry_value_added_m'],
                })
    out = pd.DataFrame(records)
    out['ebitda_margin_pct'] = out['ebitda_m'] / out['sales_m'] * 100
    out['op_profit_margin_pct'] = out['op_profit_before_tax_m'] / out['sales_m'] * 100
    out['wages_to_sales_pct'] = out['wages_m'] / out['sales_m'] * 100
    return out


def _parse_ptrs_model_sheet(ws) -> pd.DataFrame:
    records = []
    for row in ws.iter_rows(min_row=2, max_col=19, values_only=True):
        industry_code = row[0]
        industry_name = row[1]
        if not industry_code or not industry_name:
            continue

        records.append(
            {
                "anzsic_division_code": str(industry_code).strip(),
                "ptrs_industry": str(industry_name).strip(),
                "ptrs_sector_key": normalise_text(industry_name),
                "ptrs_cycle8_avg_payment_days": pd.to_numeric(row[2], errors="coerce"),
                "ptrs_cycle9_avg_payment_days": pd.to_numeric(row[3], errors="coerce"),
                "ptrs_cycle8_80th_days": pd.to_numeric(row[4], errors="coerce"),
                "ptrs_cycle9_80th_days": pd.to_numeric(row[5], errors="coerce"),
                "ptrs_cycle8_95th_days": pd.to_numeric(row[6], errors="coerce"),
                "ptrs_cycle9_95th_days": pd.to_numeric(row[7], errors="coerce"),
                "ptrs_base_ar_days": pd.to_numeric(row[8], errors="coerce"),
                "ptrs_stress_ar_days": pd.to_numeric(row[9], errors="coerce"),
                "ptrs_severe_ar_days": pd.to_numeric(row[10], errors="coerce"),
                "ptrs_conservative_multiplier": pd.to_numeric(row[11], errors="coerce"),
                "ptrs_adjusted_base_ar_days": pd.to_numeric(row[12], errors="coerce"),
                "ptrs_adjusted_stress_ar_days": pd.to_numeric(row[13], errors="coerce"),
                "ptrs_adjusted_severe_ar_days": pd.to_numeric(row[14], errors="coerce"),
                "ptrs_cycle8_paid_on_time_pct": pd.to_numeric(row[15], errors="coerce"),
                "ptrs_cycle9_paid_on_time_pct": pd.to_numeric(row[16], errors="coerce"),
                "ptrs_latest_cycle_used": row[17],
                "ptrs_model_note": row[18],
            }
        )
    return pd.DataFrame(records)


def _parse_ptrs_source_sheet(ws, cycle_number: int) -> pd.DataFrame:
    records = []
    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        industry_code = row[0]
        industry_name = row[1]
        if not industry_code or not industry_name:
            continue

        records.append(
            {
                "anzsic_division_code": str(industry_code).strip(),
                "ptrs_industry": str(industry_name).strip(),
                f"ptrs_cycle{cycle_number}_avg_common_payment_days": pd.to_numeric(row[2], errors="coerce"),
                f"ptrs_cycle{cycle_number}_avg_payment_days": pd.to_numeric(row[3], errors="coerce"),
                f"ptrs_cycle{cycle_number}_80th_days": pd.to_numeric(row[4], errors="coerce"),
                f"ptrs_cycle{cycle_number}_95th_days": pd.to_numeric(row[5], errors="coerce"),
                f"ptrs_cycle{cycle_number}_paid_on_time_pct": pd.to_numeric(row[6], errors="coerce"),
            }
        )
    return pd.DataFrame(records)


def _parse_ptrs_multiplier(ws) -> float:
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        label = row[0]
        if normalise_text(label) == "conservative multiplier":
            value = pd.to_numeric(row[1], errors="coerce")
            return float(value) if pd.notna(value) else 1.0
    return 1.0


def _build_ptrs_model_from_sources(cycle8_df: pd.DataFrame, cycle9_df: pd.DataFrame, multiplier: float) -> pd.DataFrame:
    base = cycle8_df.merge(cycle9_df, on=["anzsic_division_code", "ptrs_industry"], how="outer")
    base["ptrs_sector_key"] = base["ptrs_industry"].map(normalise_text)
    base["ptrs_conservative_multiplier"] = multiplier
    base["ptrs_base_ar_days"] = base[["ptrs_cycle8_avg_payment_days", "ptrs_cycle9_avg_payment_days"]].max(axis=1, skipna=True)
    base["ptrs_stress_ar_days"] = base[["ptrs_cycle8_80th_days", "ptrs_cycle9_80th_days"]].max(axis=1, skipna=True)
    base["ptrs_severe_ar_days"] = base[["ptrs_cycle8_95th_days", "ptrs_cycle9_95th_days"]].max(axis=1, skipna=True)
    base["ptrs_adjusted_base_ar_days"] = base["ptrs_base_ar_days"] * multiplier
    base["ptrs_adjusted_stress_ar_days"] = base["ptrs_stress_ar_days"] * multiplier
    base["ptrs_adjusted_severe_ar_days"] = base["ptrs_severe_ar_days"] * multiplier
    base["ptrs_latest_cycle_used"] = base["ptrs_cycle9_avg_payment_days"].apply(lambda value: "Cycle 9" if pd.notna(value) else "Cycle 8")
    base["ptrs_model_note"] = PTRS_MODEL_NOTE
    return base[
        [
            "anzsic_division_code",
            "ptrs_industry",
            "ptrs_sector_key",
            "ptrs_cycle8_avg_payment_days",
            "ptrs_cycle9_avg_payment_days",
            "ptrs_cycle8_80th_days",
            "ptrs_cycle9_80th_days",
            "ptrs_cycle8_95th_days",
            "ptrs_cycle9_95th_days",
            "ptrs_base_ar_days",
            "ptrs_stress_ar_days",
            "ptrs_severe_ar_days",
            "ptrs_conservative_multiplier",
            "ptrs_adjusted_base_ar_days",
            "ptrs_adjusted_stress_ar_days",
            "ptrs_adjusted_severe_ar_days",
            "ptrs_cycle8_paid_on_time_pct",
            "ptrs_cycle9_paid_on_time_pct",
            "ptrs_latest_cycle_used",
            "ptrs_model_note",
        ]
    ].copy()


def parse_ptrs_ar_workbook(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        model_df = _parse_ptrs_model_sheet(wb["Model_AR_Days"])
        if not model_df.empty and model_df["ptrs_base_ar_days"].notna().any():
            return model_df

        cycle8_df = _parse_ptrs_source_sheet(wb["PTRS_Cycle8_Official"], 8)
        cycle9_df = _parse_ptrs_source_sheet(wb["PTRS_Cycle9_Official"], 9)
        multiplier = _parse_ptrs_multiplier(wb["Assumptions"])
        return _build_ptrs_model_from_sources(cycle8_df, cycle9_df, multiplier)
    finally:
        wb.close()


from src.public_data.load_abs_manual_exports_helpers import (
    BUILDING_APPROVALS_FILENAME,
    OPTIONAL_BUILDING_ACTIVITY_FILES,
    OPTIONAL_LENDING_INDICATOR_FILES,
    REFERENCE_SEGMENTS,
    PROPERTY_ID_COLUMNS,
    load_building_approvals_reference,
    build_building_approvals_summary,
    load_optional_building_activity_extract,
    build_building_activity_summary,
    load_optional_lending_indicator_extract,
    build_housing_finance_summary,
)

__all__ = [
    "load_rba_cash_rate",
    "parse_abs_timeseries_xlsx",
    "parse_labour_force",
    "parse_building_approvals",
    "parse_australian_industry_totals",
    "parse_cpi",
    "parse_ppi",
    "parse_dwelling_approvals",
    "parse_ptrs_ar_workbook",
    "BUILDING_APPROVALS_FILENAME",
    "OPTIONAL_BUILDING_ACTIVITY_FILES",
    "OPTIONAL_LENDING_INDICATOR_FILES",
    "REFERENCE_SEGMENTS",
    "PROPERTY_ID_COLUMNS",
    "load_building_approvals_reference",
    "build_building_approvals_summary",
    "load_optional_building_activity_extract",
    "build_building_activity_summary",
    "load_optional_lending_indicator_extract",
    "build_housing_finance_summary",
]
