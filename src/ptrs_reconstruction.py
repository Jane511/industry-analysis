from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from urllib.parse import urlparse

import pandas as pd
from openpyxl import Workbook

from src.config import PUBLIC_SOURCE_URLS, PTRS_AR_WORKBOOK_FILENAME, RAW_PUBLIC_DIR_PTRS
from src.utils import normalise_text


PTRS_MODEL_NOTE = "PTRS is a proxy; override with borrower-specific AR evidence where reliable."


@dataclass(frozen=True)
class PtrsCycleDefinition:
    cycle_number: int
    period_label: str
    source_key: str
    source_table: str
    worksheet_name: str
    table_marker: str


PTRS_CYCLES = {
    8: PtrsCycleDefinition(
        cycle_number=8,
        period_label="1 Jul 2024 to 31 Dec 2024",
        source_key="ptrs_cycle_8_pdf",
        source_table="Table 10: New payment times measures by industry, Reporting Cycle 8",
        worksheet_name="PTRS_Cycle8_Official",
        table_marker="Table 10: New payment times measures by industry, Reporting Cycle 8",
    ),
    9: PtrsCycleDefinition(
        cycle_number=9,
        period_label="1 Jan 2025 to 30 Jun 2025",
        source_key="ptrs_cycle_9_pdf",
        source_table="Table 9: Payment terms and times measures by industry, Reporting Cycle 9",
        worksheet_name="PTRS_Cycle9_Official",
        table_marker="Table 9: Payment terms and times measures by industry, Reporting Cycle 9",
    ),
}


PTRS_INDUSTRIES = [
    ("A", "Agriculture, Forestry and Fishing", ["Agriculture, Forestry & Fishing"]),
    ("B", "Mining", []),
    ("C", "Manufacturing", []),
    ("D", "Electricity, Gas, Water and Waste Services", ["Electricity, Gas, Water & Waste Services"]),
    ("E", "Construction", []),
    ("F", "Wholesale Trade", []),
    ("G", "Retail Trade", []),
    ("H", "Accommodation and Food Services", ["Accommodation & Food Services"]),
    ("I", "Transport, Postal and Warehousing", ["Transport, Postal & Warehousing"]),
    ("J", "Information Media and Telecommunications", ["Information Media & Telecommunications"]),
    ("K", "Financial and Insurance Services", ["Financial & Insurance Services"]),
    ("L", "Rental, Hiring and Real Estate Services", ["Rental, Hiring & Real Estate Services"]),
    (
        "M",
        "Professional, Scientific and Technical Services",
        ["Professional, Scientific & Technical Services"],
    ),
    ("N", "Administrative and Support Services", ["Administrative & Support Services"]),
    ("O", "Public Administration and Safety", ["Public Administration & Safety"]),
    ("P", "Education and Training", ["Education & Training"]),
    ("Q", "Health Care and Social Assistance", ["Health Care & Social Assistance"]),
    ("R", "Arts and Recreation Services", ["Arts & Recreation Services"]),
    ("S", "Other Services", []),
]

PTRS_CANONICAL_ORDER = [code for code, _, _ in PTRS_INDUSTRIES]
PTRS_INDUSTRY_LOOKUP = {}
PTRS_INDUSTRY_NAMES = []
for code, official_name, aliases in PTRS_INDUSTRIES:
    for name in [official_name, *aliases]:
        PTRS_INDUSTRY_LOOKUP[normalise_text(name)] = (code, official_name)
        PTRS_INDUSTRY_NAMES.append(name)


def _industry_pattern_fragment(name: str) -> str:
    return re.escape(name).replace(r"\ ", r"\s+")


PTRS_ROW_PATTERN = re.compile(
    rf"(?P<industry>{'|'.join(sorted((_industry_pattern_fragment(name) for name in PTRS_INDUSTRY_NAMES), key=len, reverse=True))}|All\s+Industries)\s+"
    rf"(?P<common>\d+(?:\.\d+)?)\s+"
    rf"(?P<avg>\d+(?:\.\d+)?)\s+"
    rf"(?P<p80>\d+(?:\.\d+)?)\s+"
    rf"(?P<p95>\d+(?:\.\d+)?)\s+"
    rf"(?P<paid>\d+(?:\.\d+)?)%",
    re.IGNORECASE,
)


def _filename_from_url(url: str) -> str:
    return Path(urlparse(url).path).name


def _require_pypdf():
    try:
        from pypdf import PdfReader
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "pypdf is required for PTRS PDF reconstruction. Install project requirements first."
        ) from exc
    return PdfReader


def _clean_pdf_line(line: str) -> str:
    replacements = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\xa0": " ",
    }
    cleaned = str(line)
    for source, target in replacements.items():
        cleaned = cleaned.replace(source, target)
    return " ".join(cleaned.split()).strip()


def extract_pdf_text(path: Path) -> str:
    PdfReader = _require_pypdf()
    reader = PdfReader(str(path))
    return "\n".join(_clean_pdf_line(page.extract_text() or "") for page in reader.pages)


def parse_ptrs_cycle_table_from_text(text: str, cycle_number: int) -> pd.DataFrame:
    cycle = PTRS_CYCLES[cycle_number]
    marker = normalise_text(cycle.table_marker)
    lines = [_clean_pdf_line(line) for line in text.splitlines()]

    start_index = next(
        (idx for idx, line in enumerate(lines) if marker in normalise_text(line)),
        None,
    )
    if start_index is None:
        raise ValueError(f"Could not find PTRS table marker for cycle {cycle_number}.")

    table_lines = []
    for idx, line in enumerate(lines[start_index:], start=start_index):
        if idx > start_index and line.startswith("Table "):
            break
        table_lines.append(line)
    table_text = " ".join(table_lines)

    records = []
    for match in PTRS_ROW_PATTERN.finditer(table_text):
        industry_name = match.group("industry")
        industry_key = normalise_text(industry_name)
        if industry_key == "all industries":
            break

        industry_meta = PTRS_INDUSTRY_LOOKUP.get(industry_key)
        if industry_meta is None:
            continue

        code, official_name = industry_meta
        records.append(
            {
                "Industry Code": code,
                "Industry Name": official_name,
                "Avg Common Payment Term (Days)": pd.to_numeric(match.group("common"), errors="coerce"),
                "Avg Payment Time (Days)": pd.to_numeric(match.group("avg"), errors="coerce"),
                "80th pct (Days)": pd.to_numeric(match.group("p80"), errors="coerce"),
                "95th pct (Days)": pd.to_numeric(match.group("p95"), errors="coerce"),
                "Avg % Paid On Time": pd.to_numeric(match.group("paid"), errors="coerce") / 100,
                "Source URL": PUBLIC_SOURCE_URLS[cycle.source_key],
                "Source Table": cycle.source_table,
            }
        )

    df = pd.DataFrame(records)
    if df.empty:
        raise ValueError(f"No PTRS industry rows parsed for cycle {cycle_number}.")

    df = df.drop_duplicates(subset=["Industry Code"], keep="first").copy()
    df["Industry Code"] = pd.Categorical(df["Industry Code"], categories=PTRS_CANONICAL_ORDER, ordered=True)
    df = df.sort_values("Industry Code").reset_index(drop=True)
    df["Industry Code"] = df["Industry Code"].astype(str)

    if len(df) != len(PTRS_CANONICAL_ORDER):
        missing_codes = sorted(set(PTRS_CANONICAL_ORDER) - set(df["Industry Code"]))
        raise ValueError(f"PTRS cycle {cycle_number} parse incomplete. Missing industry codes: {missing_codes}")

    return df


def parse_ptrs_cycle_pdf(pdf_path: Path, cycle_number: int) -> pd.DataFrame:
    return parse_ptrs_cycle_table_from_text(extract_pdf_text(pdf_path), cycle_number)


def _set_percentage_columns(ws, columns: list[str]) -> None:
    header = {ws.cell(row=1, column=idx).value: idx for idx in range(1, ws.max_column + 1)}
    for column_name in columns:
        column_index = header.get(column_name)
        if column_index is None:
            continue
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=column_index).number_format = "0.0%"


def _source_lookup_formula(sheet_name: str, row_idx: int, return_column: str) -> str:
    return (
        f"=INDEX({sheet_name}!${return_column}$2:${return_column}$20,"
        f"MATCH($A{row_idx},{sheet_name}!$A$2:$A$20,0))"
    )


def write_ptrs_workbook(cycle8_df: pd.DataFrame, cycle9_df: pd.DataFrame, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("ReadMe")
    readme_rows = [
        ("PTRS Multi-Cycle AR Days Workbook (Official-source version)", None),
        (None, None),
        (
            "Purpose",
            "Committee-ready AR/AP benchmark workbook rebuilt automatically from official PTRS Cycle 8 and Cycle 9 industry tables.",
        ),
        (
            "What this workbook does",
            "Keeps original official data in source sheets and calculates base / stress / severe AR benchmarks in the model sheet.",
        ),
        (
            "Important scope note",
            "PTRS measures how reporting entities pay small business suppliers. It is a proxy for SME collection days, not a direct SME debtors dataset.",
        ),
        (
            "Reporting population",
            "Scheme applies to large businesses and some Commonwealth entities with annual consolidated revenue above $100m, subject to statutory criteria.",
        ),
        ("Reporting cycle cadence", "Two reporting cycles per year: 1 Jan-30 Jun and 1 Jul-31 Dec."),
        (
            "Most recent official publication used",
            "Cycle 9 (1 Jan 2025-30 Jun 2025). The January 2026 update also notes Cycle 10 has commenced but is not used here.",
        ),
    ]
    for row in readme_rows:
        readme.append(row)

    metadata = wb.create_sheet("Source_Metadata")
    metadata_rows = [
        ("Field", "Value", "Source / Notes"),
        ("Cycle 8 period", PTRS_CYCLES[8].period_label, f"Official source: {PUBLIC_SOURCE_URLS['ptrs_cycle_8_pdf']}"),
        ("Cycle 8 table", PTRS_CYCLES[8].source_table, "Used for original industry-level AR/AP proxy inputs"),
        ("Cycle 9 period", PTRS_CYCLES[9].period_label, f"Official source: {PUBLIC_SOURCE_URLS['ptrs_cycle_9_pdf']}"),
        ("Cycle 9 table", PTRS_CYCLES[9].source_table, "Used for original industry-level AR/AP proxy inputs"),
        (
            "Reporting population",
            "Reporting entities under PTRS",
            f"Guidance materials: {PUBLIC_SOURCE_URLS['ptrs_guidance']}",
        ),
        (
            "Cycle cadence",
            "Two reporting cycles per year",
            f"Guidance materials: {PUBLIC_SOURCE_URLS['ptrs_guidance']}",
        ),
        ("Cycle 10 status", "Not used in this workbook", "January 2026 update shows Cycle 10 exists but is incomplete"),
    ]
    for row in metadata_rows:
        metadata.append(row)

    sheet_map = {
        8: wb.create_sheet(PTRS_CYCLES[8].worksheet_name),
        9: wb.create_sheet(PTRS_CYCLES[9].worksheet_name),
    }
    for cycle_number, df in {8: cycle8_df, 9: cycle9_df}.items():
        ws = sheet_map[cycle_number]
        ws.append(list(df.columns))
        for row in df.itertuples(index=False, name=None):
            ws.append(row)
        _set_percentage_columns(ws, ["Avg % Paid On Time"])

    assumptions = wb.create_sheet("Assumptions")
    assumption_rows = [
        ("Assumptions & Selection Logic", None, None),
        (None, None, None),
        (
            "Conservative Multiplier",
            1.0,
            "User input. Keep at 1.00x for pure official-source base case. Increase if committee wants extra conservatism.",
        ),
        ("Base AR Selection Rule", "MAX(Cycle 8 Avg Payment Time, Cycle 9 Avg Payment Time)", None),
        ("Stress AR Selection Rule", "MAX(Cycle 8 80th pct, Cycle 9 80th pct)", None),
        ("Severe AR Selection Rule", "MAX(Cycle 8 95th pct, Cycle 9 95th pct)", None),
        ("Estimated Receivables Formula", "Annual Credit Sales / 365 x AR Days", None),
    ]
    for row in assumption_rows:
        assumptions.append(row)

    model = wb.create_sheet("Model_AR_Days")
    model_headers = [
        "Industry Code",
        "Industry Name",
        "Cycle 8 Avg Payment Time",
        "Cycle 9 Avg Payment Time",
        "Cycle 8 80th pct",
        "Cycle 9 80th pct",
        "Cycle 8 95th pct",
        "Cycle 9 95th pct",
        "Base AR Days (MAX of C8/C9 Avg)",
        "Stress AR Days (MAX of C8/C9 80th)",
        "Severe AR Days (MAX of C8/C9 95th)",
        "Conservative Multiplier",
        "Adjusted Base AR Days",
        "Adjusted Stress AR Days",
        "Adjusted Severe AR Days",
        "Cycle 8 Avg % Paid On Time",
        "Cycle 9 Avg % Paid On Time",
        "Latest Released Cycle Used?",
        "Model Note",
    ]
    model.append(model_headers)

    for model_row_idx, (industry_code, industry_name, _) in enumerate(PTRS_INDUSTRIES, start=2):
        model.append(
            [
                industry_code,
                industry_name,
                _source_lookup_formula(PTRS_CYCLES[8].worksheet_name, model_row_idx, "D"),
                _source_lookup_formula(PTRS_CYCLES[9].worksheet_name, model_row_idx, "D"),
                _source_lookup_formula(PTRS_CYCLES[8].worksheet_name, model_row_idx, "E"),
                _source_lookup_formula(PTRS_CYCLES[9].worksheet_name, model_row_idx, "E"),
                _source_lookup_formula(PTRS_CYCLES[8].worksheet_name, model_row_idx, "F"),
                _source_lookup_formula(PTRS_CYCLES[9].worksheet_name, model_row_idx, "F"),
                f"=MAX(C{model_row_idx},D{model_row_idx})",
                f"=MAX(E{model_row_idx},F{model_row_idx})",
                f"=MAX(G{model_row_idx},H{model_row_idx})",
                "=Assumptions!$B$3",
                f"=I{model_row_idx}*L{model_row_idx}",
                f"=J{model_row_idx}*L{model_row_idx}",
                f"=K{model_row_idx}*L{model_row_idx}",
                _source_lookup_formula(PTRS_CYCLES[8].worksheet_name, model_row_idx, "G"),
                _source_lookup_formula(PTRS_CYCLES[9].worksheet_name, model_row_idx, "G"),
                "Cycle 9",
                PTRS_MODEL_NOTE,
            ]
        )

    model.append([None] * len(model_headers))
    model.append([None] * len(model_headers))
    model.append(["Illustrative Estimated Receivables Block"] + [None] * (len(model_headers) - 1))
    model.append(["Selected Industry Code", "A"] + [None] * (len(model_headers) - 2))
    model.append(["Annual Credit Sales (AUD)", 10000000] + [None] * (len(model_headers) - 2))
    model.append(
        [
            "Adjusted Base AR Days",
            "=INDEX($M$2:$M$20,MATCH(B24,$A$2:$A$20,0))",
        ]
        + [None] * (len(model_headers) - 2)
    )
    model.append(
        [
            "Adjusted Stress AR Days",
            "=INDEX($N$2:$N$20,MATCH(B24,$A$2:$A$20,0))",
        ]
        + [None] * (len(model_headers) - 2)
    )
    model.append(
        [
            "Adjusted Severe AR Days",
            "=INDEX($O$2:$O$20,MATCH(B24,$A$2:$A$20,0))",
        ]
        + [None] * (len(model_headers) - 2)
    )
    model.append(["Estimated Receivables - Base", "=B25/365*B26"] + [None] * (len(model_headers) - 2))
    model.append(["Estimated Receivables - Stress", "=B25/365*B27"] + [None] * (len(model_headers) - 2))
    model.append(["Estimated Receivables - Severe", "=B25/365*B28"] + [None] * (len(model_headers) - 2))

    for ws in [sheet_map[8], sheet_map[9], model]:
        ws.freeze_panes = "A2"
    _set_percentage_columns(model, ["Cycle 8 Avg % Paid On Time", "Cycle 9 Avg % Paid On Time"])

    wb.save(output_path)
    return output_path


def reconstruct_ptrs_workbook(cycle8_pdf_path: Path, cycle9_pdf_path: Path, output_path: Path) -> Path:
    cycle8_df = parse_ptrs_cycle_pdf(cycle8_pdf_path, 8)
    cycle9_df = parse_ptrs_cycle_pdf(cycle9_pdf_path, 9)
    return write_ptrs_workbook(cycle8_df, cycle9_df, output_path)


def reconstruct_ptrs_workbook_from_downloads(download_dir: Path | None = None) -> Path:
    download_dir = download_dir or RAW_PUBLIC_DIR_PTRS
    cycle8_pdf_path = download_dir / _filename_from_url(PUBLIC_SOURCE_URLS["ptrs_cycle_8_pdf"])
    cycle9_pdf_path = download_dir / _filename_from_url(PUBLIC_SOURCE_URLS["ptrs_cycle_9_pdf"])
    output_path = download_dir / PTRS_AR_WORKBOOK_FILENAME

    if not cycle8_pdf_path.exists():
        raise FileNotFoundError(f"Missing PTRS Cycle 8 PDF: {cycle8_pdf_path}")
    if not cycle9_pdf_path.exists():
        raise FileNotFoundError(f"Missing PTRS Cycle 9 PDF: {cycle9_pdf_path}")

    return reconstruct_ptrs_workbook(cycle8_pdf_path, cycle9_pdf_path, output_path)
