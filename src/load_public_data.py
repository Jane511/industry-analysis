from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.ptrs_reconstruction import PTRS_MODEL_NOTE
from src.utils import normalise_text

def load_rba_cash_rate(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, skiprows=[0,2,3,4,5,6,7,8,9,10], encoding='utf-8-sig')
    df = df.rename(columns={'Title':'date'})
    df['date'] = pd.to_datetime(df['date'], format='%d-%b-%Y', errors='coerce')
    df['Cash Rate Target'] = pd.to_numeric(df['Cash Rate Target'], errors='coerce')
    return df[['date','Cash Rate Target']].dropna()

def parse_abs_timeseries_xlsx(path: Path, measure_name: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name='Data1', header=None)
    dates = pd.to_datetime(df.iloc[9:,0], format='mixed', errors='coerce')
    out = []
    for j in range(1, df.shape[1]):
        header = df.iloc[0, j]
        if pd.isna(header):
            continue
        industry = str(header).split(';')[2].strip() if ';' in str(header) else str(header)
        values = pd.to_numeric(df.iloc[9:,j], errors='coerce')
        tmp = pd.DataFrame({'date': dates, 'value': values}).dropna(subset=['date'])
        tmp['industry'] = industry
        tmp['measure'] = measure_name
        out.append(tmp)
    return pd.concat(out, ignore_index=True)

def parse_labour_force(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name='Data1', header=None)
    dates = pd.to_datetime(df.iloc[9:,0], format='mixed', errors='coerce')
    out = []
    for j in range(1, df.shape[1]):
        industry_raw = df.iloc[0, j]
        series_type = df.iloc[2, j]
        if pd.isna(industry_raw) or pd.isna(series_type):
            continue
        industry = str(industry_raw).split(';')[0].strip()
        values = pd.to_numeric(df.iloc[9:,j], errors='coerce')
        tmp = pd.DataFrame({'date': dates, 'value': values}).dropna(subset=['date'])
        tmp['industry'] = industry
        tmp['series_type'] = series_type
        out.append(tmp)
    return pd.concat(out, ignore_index=True)

def parse_building_approvals(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name='Data1', header=None)
    dates = pd.to_datetime(df.iloc[9:,0], format='mixed', errors='coerce')
    out = []
    for j in range(1, df.shape[1]):
        header = df.iloc[0, j]
        if pd.isna(header):
            continue
        parts = [p.strip() for p in str(header).split(';') if p.strip()]
        if len(parts) < 3:
            continue
        values = pd.to_numeric(df.iloc[9:,j], errors='coerce')
        tmp = pd.DataFrame({'date': dates, 'value': values}).dropna(subset=['date'])
        tmp['sector_group'] = parts[1]
        tmp['building_type'] = parts[2]
        out.append(tmp)
    return pd.concat(out, ignore_index=True)

def parse_australian_industry_totals(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name='Table_1', header=None)
    df.columns = ['label','employment_000','wages_m','sales_m','total_income_m','total_expenses_m','op_profit_before_tax_m','ebitda_m','industry_value_added_m']
    records = []
    for i, row in df.iterrows():
        label = row['label']
        if isinstance(label, str) and label.lower().startswith('total '):
            for k in range(1,4):
                year_row = df.iloc[i+k]
                records.append({
                    'sector': label.replace('Total ','').strip(),
                    'year': str(year_row['label']),
                    'employment_000': year_row['employment_000'],
                    'wages_m': year_row['wages_m'],
                    'sales_m': year_row['sales_m'],
                    'op_profit_before_tax_m': year_row['op_profit_before_tax_m'],
                    'ebitda_m': year_row['ebitda_m'],
                    'industry_value_added_m': year_row['industry_value_added_m'],
                })
    out = pd.DataFrame(records)
    out['ebitda_margin_pct'] = out['ebitda_m'] / out['sales_m'] * 100
    out['op_profit_margin_pct'] = out['op_profit_before_tax_m'] / out['sales_m'] * 100
    out['wages_to_sales_pct'] = out['wages_m'] / out['sales_m'] * 100
    return out


def _parse_ptrs_model_sheet(ws) -> pd.DataFrame:
    records = []
    for row in ws.iter_rows(min_row=2, max_col=19, values_only=True):
        industry_code = row[0]
        industry_name = row[1]
        if not industry_code or not industry_name:
            continue

        records.append(
            {
                "anzsic_division_code": str(industry_code).strip(),
                "ptrs_industry": str(industry_name).strip(),
                "ptrs_sector_key": normalise_text(industry_name),
                "ptrs_cycle8_avg_payment_days": pd.to_numeric(row[2], errors="coerce"),
                "ptrs_cycle9_avg_payment_days": pd.to_numeric(row[3], errors="coerce"),
                "ptrs_cycle8_80th_days": pd.to_numeric(row[4], errors="coerce"),
                "ptrs_cycle9_80th_days": pd.to_numeric(row[5], errors="coerce"),
                "ptrs_cycle8_95th_days": pd.to_numeric(row[6], errors="coerce"),
                "ptrs_cycle9_95th_days": pd.to_numeric(row[7], errors="coerce"),
                "ptrs_base_ar_days": pd.to_numeric(row[8], errors="coerce"),
                "ptrs_stress_ar_days": pd.to_numeric(row[9], errors="coerce"),
                "ptrs_severe_ar_days": pd.to_numeric(row[10], errors="coerce"),
                "ptrs_conservative_multiplier": pd.to_numeric(row[11], errors="coerce"),
                "ptrs_adjusted_base_ar_days": pd.to_numeric(row[12], errors="coerce"),
                "ptrs_adjusted_stress_ar_days": pd.to_numeric(row[13], errors="coerce"),
                "ptrs_adjusted_severe_ar_days": pd.to_numeric(row[14], errors="coerce"),
                "ptrs_cycle8_paid_on_time_pct": pd.to_numeric(row[15], errors="coerce"),
                "ptrs_cycle9_paid_on_time_pct": pd.to_numeric(row[16], errors="coerce"),
                "ptrs_latest_cycle_used": row[17],
                "ptrs_model_note": row[18],
            }
        )
    return pd.DataFrame(records)


def _parse_ptrs_source_sheet(ws, cycle_number: int) -> pd.DataFrame:
    records = []
    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        industry_code = row[0]
        industry_name = row[1]
        if not industry_code or not industry_name:
            continue

        records.append(
            {
                "anzsic_division_code": str(industry_code).strip(),
                "ptrs_industry": str(industry_name).strip(),
                f"ptrs_cycle{cycle_number}_avg_common_payment_days": pd.to_numeric(row[2], errors="coerce"),
                f"ptrs_cycle{cycle_number}_avg_payment_days": pd.to_numeric(row[3], errors="coerce"),
                f"ptrs_cycle{cycle_number}_80th_days": pd.to_numeric(row[4], errors="coerce"),
                f"ptrs_cycle{cycle_number}_95th_days": pd.to_numeric(row[5], errors="coerce"),
                f"ptrs_cycle{cycle_number}_paid_on_time_pct": pd.to_numeric(row[6], errors="coerce"),
            }
        )
    return pd.DataFrame(records)


def _parse_ptrs_multiplier(ws) -> float:
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        label = row[0]
        if normalise_text(label) == "conservative multiplier":
            value = pd.to_numeric(row[1], errors="coerce")
            return float(value) if pd.notna(value) else 1.0
    return 1.0


def _build_ptrs_model_from_sources(cycle8_df: pd.DataFrame, cycle9_df: pd.DataFrame, multiplier: float) -> pd.DataFrame:
    base = cycle8_df.merge(cycle9_df, on=["anzsic_division_code", "ptrs_industry"], how="outer")
    base["ptrs_sector_key"] = base["ptrs_industry"].map(normalise_text)
    base["ptrs_conservative_multiplier"] = multiplier
    base["ptrs_base_ar_days"] = base[["ptrs_cycle8_avg_payment_days", "ptrs_cycle9_avg_payment_days"]].max(axis=1, skipna=True)
    base["ptrs_stress_ar_days"] = base[["ptrs_cycle8_80th_days", "ptrs_cycle9_80th_days"]].max(axis=1, skipna=True)
    base["ptrs_severe_ar_days"] = base[["ptrs_cycle8_95th_days", "ptrs_cycle9_95th_days"]].max(axis=1, skipna=True)
    base["ptrs_adjusted_base_ar_days"] = base["ptrs_base_ar_days"] * multiplier
    base["ptrs_adjusted_stress_ar_days"] = base["ptrs_stress_ar_days"] * multiplier
    base["ptrs_adjusted_severe_ar_days"] = base["ptrs_severe_ar_days"] * multiplier
    base["ptrs_latest_cycle_used"] = base["ptrs_cycle9_avg_payment_days"].apply(lambda value: "Cycle 9" if pd.notna(value) else "Cycle 8")
    base["ptrs_model_note"] = PTRS_MODEL_NOTE
    return base[
        [
            "anzsic_division_code",
            "ptrs_industry",
            "ptrs_sector_key",
            "ptrs_cycle8_avg_payment_days",
            "ptrs_cycle9_avg_payment_days",
            "ptrs_cycle8_80th_days",
            "ptrs_cycle9_80th_days",
            "ptrs_cycle8_95th_days",
            "ptrs_cycle9_95th_days",
            "ptrs_base_ar_days",
            "ptrs_stress_ar_days",
            "ptrs_severe_ar_days",
            "ptrs_conservative_multiplier",
            "ptrs_adjusted_base_ar_days",
            "ptrs_adjusted_stress_ar_days",
            "ptrs_adjusted_severe_ar_days",
            "ptrs_cycle8_paid_on_time_pct",
            "ptrs_cycle9_paid_on_time_pct",
            "ptrs_latest_cycle_used",
            "ptrs_model_note",
        ]
    ].copy()


def parse_ptrs_ar_workbook(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        model_df = _parse_ptrs_model_sheet(wb["Model_AR_Days"])
        if not model_df.empty and model_df["ptrs_base_ar_days"].notna().any():
            return model_df

        cycle8_df = _parse_ptrs_source_sheet(wb["PTRS_Cycle8_Official"], 8)
        cycle9_df = _parse_ptrs_source_sheet(wb["PTRS_Cycle9_Official"], 9)
        multiplier = _parse_ptrs_multiplier(wb["Assumptions"])
        return _build_ptrs_model_from_sources(cycle8_df, cycle9_df, multiplier)
    finally:
        wb.close()
